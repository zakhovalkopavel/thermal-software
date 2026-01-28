#!/usr/bin/env python3
"""
Thermophysical Data Aggregation & Normalization for Ionic Compounds
Step-by-step workflow with user confirmation gates

This script processes NBS thermodynamic tables and produces standardized
CSV outputs with thermophysical properties for specified ionic compounds.
"""

import pandas as pd
import numpy as np
import openpyxl
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import sys


class ThermophysicalDataProcessor:
    """Main processor for thermophysical data aggregation"""

    # Target cations
    TARGET_CATIONS = ['Na', 'K', 'Li', 'Mg', 'Ca', 'Ba', 'Fe', 'Zn', 'Al', 'Ti', 'Si', 'B']

    # Target anion families with search patterns
    ANION_FAMILIES = {
        'chloride': {
            'patterns': [r'Cl(?:\d+)?(?:\s|$|,)', r'chloride'],
            'elements': ['Cl']
        },
        'fluoride': {
            'patterns': [r'F(?:\d+)?(?:\s|$|,)', r'fluoride'],
            'elements': ['F']
        },
        'carbonate': {
            'patterns': [r'CO3', r'C\s*O3', r'carbonate'],
            'elements': ['C', 'O']
        },
        'sulfate': {
            'patterns': [r'SO4', r'S\s*O4', r'sulfate'],
            'elements': ['S', 'O']
        },
        'borate': {
            'patterns': [r'B\d*O\d+', r'borate', r'tetraborate'],
            'elements': ['B', 'O']
        },
        'tetraborate': {
            'patterns': [r'B4O7', r'tetraborate'],
            'elements': ['B', 'O']
        },
        'phosphate': {
            'patterns': [r'P\d*O\d+', r'phosphate', r'metaphosphate'],
            'elements': ['P', 'O']
        },
        'metaphosphate': {
            'patterns': [r'PO3', r'metaphosphate'],
            'elements': ['P', 'O']
        },
        'oxide': {
            'patterns': [r'O\d*(?:\s|$|,)(?!x)', r'oxide'],
            'elements': ['O']
        }
    }

    def __init__(self, input_file: str):
        """Initialize processor with input file path"""
        self.input_file = input_file
        self.df_raw = None
        self.df_extracted = None
        self.df_converted = None
        self.df_enriched = None
        self.df_final = None

    def load_nbs_data(self) -> pd.DataFrame:
        """Load NBS Tables from Excel file"""
        print("Loading NBS Tables Library...")

        # Load workbook
        wb = openpyxl.load_workbook(self.input_file, read_only=True)
        ws = wb['NBS Tables']

        # Define column names based on NBS table structure
        columns = [
            'Formula',
            'Solvent',
            'Name',
            'State_Description',
            'State',
            'Molar_Mass_g_mol',
            'DfH0_kJ_mol',
            'DfH_298K_kJ_mol',
            'DfG_298K_kJ_mol',
            'H_H0_298K_kJ_mol',
            'S_298K_J_mol_K',
            'Cp_298K_J_mol_K',
            'Extra'
        ]

        # Extract data (skip first 2 header rows)
        data = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            data.append(row)

        # Create DataFrame
        df = pd.DataFrame(data, columns=columns)

        print(f"Loaded {len(df)} rows from NBS Tables")
        return df

    def identify_cation(self, formula: str) -> Optional[str]:
        """Identify the primary cation in a formula"""
        if not formula or pd.isna(formula) or not isinstance(formula, str):
            return None

        # Try to match each target cation at the start of the formula
        for cation in self.TARGET_CATIONS:
            # Match cation at start, possibly followed by numbers or non-letters
            pattern = f'^{re.escape(cation)}(?:[^a-z]|$)'
            if re.search(pattern, formula, re.IGNORECASE):
                return cation

        return None

    def identify_anion_family(self, formula: str, name: str) -> Optional[str]:
        """Identify the anion family from formula and name"""
        if not formula:
            return None

        formula_str = str(formula).strip()
        name_str = str(name).strip().lower() if name and not pd.isna(name) else ""

        # Check each anion family
        for family, info in self.ANION_FAMILIES.items():
            # Check formula patterns
            for pattern in info['patterns']:
                if re.search(pattern, formula_str, re.IGNORECASE) or \
                   re.search(pattern, name_str, re.IGNORECASE):
                    return family

        return None

    def parse_hydration(self, formula: str, state_desc: str) -> Tuple[str, Optional[str]]:
        """Parse hydration state from formula and state description"""
        if not formula:
            return "unknown", None

        formula_str = str(formula).strip()
        state_str = str(state_desc).lower() if state_desc and not pd.isna(state_desc) else ""

        # Check for hydration patterns
        hydrate_match = re.search(r'(\d+)H2O|(\d+)H₂O|hydrate|monohydrate|dihydrate|trihydrate',
                                  formula_str + " " + state_str, re.IGNORECASE)

        if hydrate_match:
            if 'mono' in state_str.lower() or '1H2O' in formula_str or '1H₂O' in formula_str:
                return "monohydrate", "H2O"
            elif 'di' in state_str.lower() or '2H2O' in formula_str or '2H₂O' in formula_str:
                return "dihydrate", "2H2O"
            elif 'tri' in state_str.lower() or '3H2O' in formula_str or '3H₂O' in formula_str:
                return "trihydrate", "3H2O"
            else:
                num = hydrate_match.group(1) or hydrate_match.group(2)
                if num:
                    return f"{num}-hydrate", f"{num}H2O"
                return "hydrate", None

        if 'anhydrous' in state_str:
            return "anhydrous", None

        # If no hydration info, assume anhydrous for solid compounds
        return "anhydrous", None

    def infer_oxidation_state(self, formula: str) -> Optional[str]:
        """Infer oxidation state for multivalent cations"""
        if not formula:
            return None

        # Look for explicit oxidation state notation
        ox_match = re.search(r'\(([IV]+)\)', formula)
        if ox_match:
            return ox_match.group(1)

        # Infer from subscripts for Fe, Ti
        if 'Fe' in formula:
            if 'Fe2' in formula or 'FeO' in formula:
                return 'II'
            elif 'Fe3' in formula or 'Fe2O3' in formula:
                return 'III'

        if 'Ti' in formula:
            if 'TiO2' in formula:
                return 'IV'
            elif 'Ti2O3' in formula:
                return 'III'

        return None

    def step1_extract_input_only(self) -> pd.DataFrame:
        """
        STEP 1: Ingest & Extract (Input-Only)
        Filter NBS data to target compounds without adding external data
        """
        print("\n" + "="*80)
        print("STEP 1: INGEST & EXTRACT (INPUT-ONLY)")
        print("="*80)

        # Load raw data
        self.df_raw = self.load_nbs_data()

        # Filter out section headers and invalid rows
        df = self.df_raw.copy()
        df = df[df['Formula'].notna()]
        df = df[~df['Formula'].astype(str).str.contains('^Table', na=False, regex=True)]
        df = df[df['State'].notna()]

        print(f"\nAfter removing headers: {len(df)} rows")

        # Identify compounds with target cations
        df['cation'] = df['Formula'].apply(self.identify_cation)
        df_cations = df[df['cation'].notna()].copy()
        print(f"Rows with target cations: {len(df_cations)}")

        # Identify anion families
        df_cations['anion_family'] = df_cations.apply(
            lambda row: self.identify_anion_family(row['Formula'], row['Name']),
            axis=1
        )
        df_target = df_cations[df_cations['anion_family'].notna()].copy()
        print(f"Rows matching target anion families: {len(df_target)}")

        # Parse additional metadata
        df_target['hydration'], df_target['hydration_formula'] = zip(*df_target.apply(
            lambda row: self.parse_hydration(row['Formula'], row['State_Description']),
            axis=1
        ))

        df_target['oxidation_state'] = df_target['Formula'].apply(self.infer_oxidation_state)

        # Standardize state representation
        df_target['phase_at_25C'] = df_target['State'].map({
            'cr': 'solid',
            'c': 'solid',
            'l': 'liquid',
            'g': 'gas',
            'aq': 'aqueous',
            'ao': 'aqueous',
            'am': 'amorphous'
        })

        # Create standardized output columns
        output_df = pd.DataFrame({
            # Core identification
            'cation': df_target['cation'],
            'anion_family': df_target['anion_family'],
            'compound_name': df_target['Name'],
            'formula': df_target['Formula'],
            'oxidation_state': df_target['oxidation_state'],
            'hydration': df_target['hydration'],
            'hydration_formula': df_target['hydration_formula'],
            'phase_at_25C': df_target['phase_at_25C'],

            # Thermal properties (from NBS - molar basis)
            'Cp_molar_J_per_molK_298K': df_target['Cp_298K_J_mol_K'],
            'Cp_ref_temperature_K': 298.15,  # NBS data is at 298.15 K
            'Cp_mass_J_per_kgK_298K': None,  # To be computed in Step 2

            # Temperature properties (NBS doesn't have Tm, Tb, Td directly)
            'Tm_C': None,
            'Tm_K': None,
            'Tb_C': None,
            'Tb_K': None,
            'Td_C': None,

            # Provenance
            'pressure_kPa': 101.325,  # NBS standard pressure
            'source_Cp': 'NBS Tables',
            'source_Cp_link': None,
            'source_T_transition': None,
            'source_T_transition_link': None,
            'data_quality': 'A',  # NBS is authoritative primary source
            'CAS': None,  # Not in NBS tables
            'molar_mass_g_per_mol': df_target['Molar_Mass_g_mol'],
            'notes': None,

            # Processing metadata
            'conversion_basis': None,
            'value_status': 'original'
        })

        # Clean up numeric conversions
        output_df['Cp_molar_J_per_molK_298K'] = pd.to_numeric(
            output_df['Cp_molar_J_per_molK_298K'], errors='coerce'
        )
        output_df['molar_mass_g_per_mol'] = pd.to_numeric(
            output_df['molar_mass_g_per_mol'], errors='coerce'
        )

        # Filter to only solid/liquid phases (exclude gas and aqueous for now)
        output_df = output_df[output_df['phase_at_25C'].isin(['solid', 'liquid', 'amorphous'])]

        print(f"\nFinal extracted compounds: {len(output_df)}")
        print("\nBreakdown by cation:")
        print(output_df['cation'].value_counts().sort_index())
        print("\nBreakdown by anion family:")
        print(output_df['anion_family'].value_counts().sort_index())

        self.df_extracted = output_df
        return output_df

    def preview_step1_results(self, n: int = 20):
        """Display preview of Step 1 results"""
        if self.df_extracted is None:
            print("No extracted data available. Run step1_extract_input_only() first.")
            return

        print("\n" + "="*80)
        print("STEP 1 PREVIEW - First 20 compounds")
        print("="*80)

        preview_cols = [
            'cation', 'anion_family', 'formula', 'compound_name',
            'phase_at_25C', 'Cp_molar_J_per_molK_298K', 'molar_mass_g_per_mol'
        ]

        print(self.df_extracted[preview_cols].head(n).to_string(index=False))

        print(f"\nTotal compounds: {len(self.df_extracted)}")
        print(f"Compounds with Cp data: {self.df_extracted['Cp_molar_J_per_molK_298K'].notna().sum()}")

    def save_step1_csv(self, output_path: str = None):
        """Save Step 1 results to CSV"""
        if self.df_extracted is None:
            print("No extracted data to save.")
            return

        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d')
            output_path = f'thermophysical_data_{timestamp}_step1_input_only.csv'

        self.df_extracted.to_csv(output_path, index=False, encoding='utf-8')
        print(f"\nStep 1 output saved to: {output_path}")
        return output_path


def main():
    """Main execution function"""

    # Initialize processor
    input_file = '/opt/thermal-software/library/NBS_Tables Library.xlsx'
    processor = ThermophysicalDataProcessor(input_file)

    # STEP 1: Extract input-only data
    print("\n" + "#"*80)
    print("# THERMOPHYSICAL DATA AGGREGATION - STEP 1")
    print("#"*80)

    df_step1 = processor.step1_extract_input_only()
    processor.preview_step1_results(n=30)

    # Save Step 1 output
    output_file = processor.save_step1_csv()

    print("\n" + "="*80)
    print("STEP 1 COMPLETE")
    print("="*80)
    print("\nNext: Review the output and confirm to proceed to Step 2")
    print("Step 2 will add kg-based Cp columns and unified temperature units")


if __name__ == '__main__':
    main()

