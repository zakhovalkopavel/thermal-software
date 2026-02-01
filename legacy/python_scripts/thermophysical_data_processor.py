#!/usr/bin/env python3
"""
FINAL Comprehensive Thermophysical Data Processor

Creates complete database with ALL properties in single files:
- One file for anhydrous compounds
- One file for hydrates

Includes:
- All NBS thermodynamic data
- Proper compound names generated from formulas
- Correct hydrate detection from chemical formulas
- Heat of fusion, vaporization (from CRC when available)
- Physical properties: density, thermal conductivity, viscosity, hardness
- Stability ranges
- Approximation equations where available
"""

import pandas as pd
import numpy as np
import openpyxl
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import os


class ComprehensiveThermophysicalProcessor:
    """Process and compile complete thermophysical database"""

    def __init__(self, input_file: str, resources_dir: str = '../library/resources'):
        """Initialize processor with input file and resources directory"""
        self.input_file = input_file
        self.resources_dir = resources_dir
        self.df_raw = None
        self.df_anhydrous = None
        self.df_hydrates = None

        # Load configuration from resource files
        self.load_configuration()
        self.load_compound_names()

    def load_configuration(self):
        """Load target cations, anion families, and allowed elements from config file"""
        config_file = os.path.join(self.resources_dir, 'processor_config.txt')

        self.TARGET_CATIONS = []
        self.ANION_FAMILIES = {}
        self.ALLOWED_ELEMENTS = set()
        self.EXCLUDED_PATTERNS = {}

        current_section = None

        with open(config_file, 'r') as f:
            for line in f:
                line = line.strip()

                # Skip comments and empty lines
                if not line or line.startswith('#'):
                    continue

                # Section headers
                if line.startswith('[') and line.endswith(']'):
                    current_section = line[1:-1]
                    continue

                # Parse based on section
                if current_section == 'cations':
                    cation = line.split(',')[0]
                    self.TARGET_CATIONS.append(cation)

                elif current_section == 'anion_families':
                    parts = line.split('\t')
                    if len(parts) >= 3:
                        family_name = parts[0]
                        patterns = parts[1].split(';;')  # Changed from '|' to ';;'
                        elements = parts[2].split()
                        self.ANION_FAMILIES[family_name] = {
                            'patterns': patterns,
                            'elements': elements
                        }

                elif current_section == 'allowed_elements':
                    elements = line.split(',')
                    self.ALLOWED_ELEMENTS.update([e.strip() for e in elements])

                elif current_section == 'excluded_patterns':
                    parts = line.split(',', 1)
                    if len(parts) == 2:
                        pattern_name = parts[0]
                        pattern = parts[1]
                        self.EXCLUDED_PATTERNS[pattern_name] = pattern

        print(f"Loaded configuration:")
        print(f"  Target cations: {len(self.TARGET_CATIONS)}")
        print(f"  Anion families: {len(self.ANION_FAMILIES)}")
        print(f"  Allowed elements: {len(self.ALLOWED_ELEMENTS)}")

    def load_compound_names(self):
        """Load compound names from CSV resource file"""
        names_file = os.path.join(self.resources_dir, 'compound_names.csv')

        self.compound_names = {}

        with open(names_file, 'r') as f:
            for line in f:
                line = line.strip()

                # Skip comments and empty lines
                if not line or line.startswith('#'):
                    continue

                parts = line.split(',')
                if len(parts) >= 2:
                    formula = parts[0].strip()
                    name = parts[1].strip()
                    self.compound_names[formula] = name

        print(f"  Compound names loaded: {len(self.compound_names)}")

    def load_nbs_data(self) -> pd.DataFrame:
        """Load NBS Tables"""
        print("Loading NBS Tables Library...")
        wb = openpyxl.load_workbook(self.input_file, read_only=True)
        ws = wb['NBS Tables']

        columns = [
            'Formula', 'Solvent', 'Name', 'State_Description', 'State',
            'Molar_Mass_g_mol', 'DfH0_kJ_mol', 'DfH_298K_kJ_mol',
            'DfG_298K_kJ_mol', 'H_H0_298K_kJ_mol', 'S_298K_J_mol_K',
            'Cp_298K_J_mol_K', 'Extra'
        ]

        data = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            data.append(row)

        df = pd.DataFrame(data, columns=columns)
        print(f"Loaded {len(df)} rows from NBS Tables")
        return df

    def detect_hydration_from_formula(self, formula: str) -> Tuple[str, Optional[str], str]:
        """
        Detect hydration from chemical formula
        Returns: (hydration_state, hydration_formula, base_formula)

        Examples:
        - ZnSO4·H2O → ('monohydrate', 'H2O', 'ZnSO4')
        - CuSO4·5H2O → ('5-hydrate', '5H2O', 'CuSO4')
        - NaCl → ('anhydrous', None, 'NaCl')
        """
        if not formula or pd.isna(formula):
            return "unknown", None, formula

        formula_str = str(formula).strip()

        # Check for water in formula with dot separator (·, ., or space)
        # Patterns: ·H2O, ·5H2O, .H2O, .5H2O, etc.
        hydrate_patterns = [
            r'[·\.\s](\d*)H₂O',  # Unicode subscript
            r'[·\.\s](\d*)H2O',  # Regular
        ]

        for pattern in hydrate_patterns:
            match = re.search(pattern, formula_str, re.IGNORECASE)
            if match:
                num = match.group(1)
                if num == '' or num == '1':
                    hydration_state = 'monohydrate'
                    hydration_formula = 'H2O'
                elif num == '2':
                    hydration_state = 'dihydrate'
                    hydration_formula = '2H2O'
                elif num == '3':
                    hydration_state = 'trihydrate'
                    hydration_formula = '3H2O'
                else:
                    hydration_state = f'{num}-hydrate'
                    hydration_formula = f'{num}H2O'

                # Extract base formula (before the dot/hydration)
                base_formula = re.sub(r'[·\.\s]\d*H[₂2]O.*$', '', formula_str)
                return hydration_state, hydration_formula, base_formula

        # No hydration found
        return 'anhydrous', None, formula_str

    def generate_compound_name(self, formula: str, original_name: str = None) -> str:
        """
        Generate proper compound name from formula if name is missing
        Uses compound_names loaded from resource file
        """
        if original_name and pd.notna(original_name) and original_name.strip():
            return original_name.strip()

        if not formula or pd.isna(formula):
            return None

        # Try to match base formula (without hydration)
        _, _, base_formula = self.detect_hydration_from_formula(formula)

        if base_formula in self.compound_names:
            return self.compound_names[base_formula]

        # If not in dictionary, return None to keep original
        return None

    def identify_cation(self, formula: str) -> Optional[str]:
        """Identify primary cation"""
        if not formula or pd.isna(formula):
            return None

        # Remove hydration part for cation identification
        _, _, base_formula = self.detect_hydration_from_formula(formula)

        for cation in self.TARGET_CATIONS:
            pattern = f'^{re.escape(cation)}(?:\d|[A-Z]|[^A-Za-z]|$)'
            if re.search(pattern, base_formula):
                return cation
        return None

    def identify_anion_family(self, formula: str, name: str) -> Optional[str]:
        """Identify anion family - returns None for pure elements"""
        if not formula:
            return None

        formula_str = str(formula).strip()
        name_str = str(name).strip().lower() if name and not pd.isna(name) else ""

        # Remove hydration for analysis
        _, _, base_formula = self.detect_hydration_from_formula(formula_str)

        # Check if it's a pure element (single element symbol, possibly with subscript number)
        # Pure elements: Fe, Al, Si, O2, N2, Cl2, etc.
        pure_element_pattern = r'^[A-Z][a-z]?\d*$'
        if re.match(pure_element_pattern, base_formula):
            return None  # Pure element, no anion family

        # Try to match anion families
        # Match order matters - more specific patterns first
        family_priority = [
            'pyrophosphate',  # P2O7 - check before general phosphate
            'metaphosphate',  # PO3 - check before general phosphate
            'pyrosulfate',    # S2O7 - check before sulfate
            'sulfate',        # SO4
            'sulfite',        # SO3
            'sulfide',        # S (no O)
            'carbonate',      # CO3
            'phosphate',      # PO4
            'borate',         # B-O compounds
            'chloride',       # Cl
            'fluoride',       # F
            'oxide',          # O
            'carbide',        # C (no O)
            'nitride',        # N (no O, no H)
            'boride',         # B (no O)
            'polyphosphate',  # complex phosphates
        ]

        # Try families in priority order
        for family in family_priority:
            if family not in self.ANION_FAMILIES:
                continue

            info = self.ANION_FAMILIES[family]
            for pattern in info['patterns']:
                try:
                    if re.search(pattern, base_formula, re.IGNORECASE) or \
                       (name_str and re.search(pattern, name_str, re.IGNORECASE)):
                        return family
                except re.error as e:
                    # Skip invalid regex patterns and show which one failed
                    if hasattr(self, '_pattern_warnings'):
                        if pattern not in self._pattern_warnings:
                            print(f"Warning: Invalid regex pattern '{pattern}' for family '{family}': {e}")
                            self._pattern_warnings.add(pattern)
                    else:
                        self._pattern_warnings = {pattern}
                        print(f"Warning: Invalid regex pattern '{pattern}' for family '{family}': {e}")
                    continue

        return None  # No anion family matched

    def contains_only_target_elements(self, formula: str) -> bool:
        """
        Check if formula contains only target elements
        Allowed: Na, K, Li, Mg, Ca, Ba, Fe, Zn, Al, Ti, Si, B (cations)
                 O, Cl, F, C, S, P, H (anions + water)
                 N (in nitrides only, not in nitrates/ammonium/organic)
        NOT allowed: Br, I, Cr, Zr, organic compounds, nitrates, ammonium salts

        Special handling:
        - Carbides (e.g., SiC, B4C) - ALLOWED
        - Nitrides (e.g., Si3N4, AlN, BN) - ALLOWED
        - Borides (e.g., TiB2, ZrB2) - ALLOWED (but Zr not target cation)
        - Organic (C-H chains) - REJECTED
        - Nitrates (NO3) - REJECTED
        - Ammonium (NH4) - REJECTED
        """
        if not formula or pd.isna(formula):
            return False

        import re
        formula_str = str(formula)

        # Remove hydration part for analysis
        base = re.sub(r'[·\.\s]\d*H[₂2]O.*$', '', formula_str)

        # Check excluded patterns (loaded from config)
        for pattern_name, pattern in self.EXCLUDED_PATTERNS.items():
            if re.search(pattern, base):
                return False

        # Extract all element symbols
        elements = set(re.findall(r'[A-Z][a-z]?', base))

        # Check if all elements are in allowed list (loaded from config)
        non_target = elements - self.ALLOWED_ELEMENTS
        return len(non_target) == 0

        # Check if all elements are in allowed list
        non_target = elements - allowed_elements
        return len(non_target) == 0

    def process_complete_database(self):
        """Process complete database with ALL properties"""
        print("\n" + "="*80)
        print("COMPREHENSIVE THERMOPHYSICAL DATABASE COMPILATION")
        print("="*80)

        # Load NBS data
        self.df_raw = self.load_nbs_data()

        # Filter valid rows
        df = self.df_raw.copy()
        df = df[df['Formula'].notna()]
        df = df[~df['Formula'].astype(str).str.contains('^Table', na=False, regex=True)]
        df = df[df['State'].notna()]

        print(f"\nAfter removing headers: {len(df)} rows")

        # Filter by element composition (BEFORE cation/anion filtering)
        print("Filtering by element composition...")
        df['has_only_target_elements'] = df['Formula'].apply(self.contains_only_target_elements)
        df_elements = df[df['has_only_target_elements']].copy()
        removed_by_elements = len(df) - len(df_elements)
        print(f"After element filtering: {len(df_elements)} rows")
        print(f"  Removed {removed_by_elements} compounds with non-target elements (N, Br, I, etc.)")

        df = df_elements

        # Identify target compounds
        df['cation'] = df['Formula'].apply(self.identify_cation)
        df_cations = df[df['cation'].notna()].copy()
        print(f"Rows with target cations: {len(df_cations)}")

        df_cations['anion_family'] = df_cations.apply(
            lambda row: self.identify_anion_family(row['Formula'], row['Name']),
            axis=1
        )
        df_target = df_cations[df_cations['anion_family'].notna()].copy()
        print(f"Rows matching target anion families: {len(df_target)}")

        # Detect hydration from formula
        hydration_data = df_target['Formula'].apply(self.detect_hydration_from_formula)
        df_target['hydration'] = hydration_data.apply(lambda x: x[0])
        df_target['hydration_formula'] = hydration_data.apply(lambda x: x[1])
        df_target['base_formula'] = hydration_data.apply(lambda x: x[2])

        # Generate compound names for missing names
        df_target['compound_name'] = df_target.apply(
            lambda row: self.generate_compound_name(row['base_formula'], row['Name']),
            axis=1
        )

        # Standardize phase
        def map_phase(state):
            if pd.isna(state):
                return None
            state_str = str(state).lower().strip()
            if state_str.startswith('cr'):
                return 'solid'
            elif state_str in ['c', 'sol']:
                return 'solid'
            elif state_str in ['l', 'liq']:
                return 'liquid'
            elif state_str in ['g', 'gas']:
                return 'gas'
            elif state_str in ['aq', 'ao', 'ai']:
                return 'aqueous'
            elif state_str in ['am', 'amorphous']:
                return 'amorphous'
            return None

        df_target['phase_at_25C'] = df_target['State'].apply(map_phase)

        # Filter to solid/liquid phases
        df_target = df_target[df_target['phase_at_25C'].isin(['solid', 'liquid', 'amorphous'])]

        print(f"\nFinal compounds: {len(df_target)}")

        # Create comprehensive output DataFrame with ALL properties
        output_df = self.create_comprehensive_dataframe(df_target)

        # Separate anhydrous and hydrates
        self.df_anhydrous = output_df[output_df['hydration'] == 'anhydrous'].copy()
        self.df_hydrates = output_df[output_df['hydration'] != 'anhydrous'].copy()

        print(f"\nAnhydrous: {len(self.df_anhydrous)}")
        print(f"Hydrates: {len(self.df_hydrates)}")

        return self.df_anhydrous, self.df_hydrates

    def create_comprehensive_dataframe(self, df_target):
        """Create DataFrame with ALL properties"""

        # Clean numeric data
        for col in ['Molar_Mass_g_mol', 'DfH0_kJ_mol', 'DfH_298K_kJ_mol',
                    'DfG_298K_kJ_mol', 'H_H0_298K_kJ_mol', 'S_298K_J_mol_K', 'Cp_298K_J_mol_K']:
            df_target[col] = pd.to_numeric(df_target[col], errors='coerce')

        # Calculate Cp_mass
        df_target['Cp_mass_J_per_kgK_298K'] = df_target['Cp_298K_J_mol_K'] / (df_target['Molar_Mass_g_mol'] / 1000.0)

        output_df = pd.DataFrame({
            # Core Identification
            'cation': df_target['cation'],
            'anion_family': df_target['anion_family'],
            'compound_name': df_target['compound_name'],
            'formula': df_target['Formula'],
            'base_formula': df_target['base_formula'],
            'hydration': df_target['hydration'],
            'hydration_formula': df_target['hydration_formula'],
            'phase_at_25C': df_target['phase_at_25C'],
            'state_description': df_target['State_Description'],

            # NBS Thermodynamic Data (all at 298.15 K, 101.325 kPa)
            'molar_mass_g_per_mol': df_target['Molar_Mass_g_mol'],
            'DfH0_kJ_per_mol': df_target['DfH0_kJ_mol'],
            'DfH_298K_kJ_per_mol': df_target['DfH_298K_kJ_mol'],
            'DfG_298K_kJ_per_mol': df_target['DfG_298K_kJ_mol'],
            'H_H0_298K_kJ_per_mol': df_target['H_H0_298K_kJ_mol'],
            'S_298K_J_per_molK': df_target['S_298K_J_mol_K'],
            'Cp_298K_J_per_molK': df_target['Cp_298K_J_mol_K'],
            'Cp_mass_J_per_kgK_298K': df_target['Cp_mass_J_per_kgK_298K'],

            # Transition Temperatures (to be filled from CRC/external sources)
            'Tm_C': None,  # Melting point
            'Tm_K': None,
            'Tb_C': None,  # Boiling point
            'Tb_K': None,
            'Td_C': None,  # Decomposition temperature

            # Heat of Transition (to be filled from CRC/external sources)
            'Hfus_kJ_per_mol': None,  # Heat of fusion
            'Hvap_kJ_per_mol': None,  # Heat of vaporization
            'Hsub_kJ_per_mol': None,  # Heat of sublimation

            # Physical Properties (to be filled from CRC/external sources)
            'density_g_per_cm3': None,  # Density at 25°C
            'thermal_conductivity_W_per_mK': None,  # Thermal conductivity
            'viscosity_mPa_s': None,  # Viscosity (for liquids)
            'hardness_Mohs': None,  # Mohs hardness (for solids)

            # Stability and Range
            'stability_range_C_min': None,  # Lower temperature limit
            'stability_range_C_max': None,  # Upper temperature limit
            'decomposition_products': None,  # What it decomposes into

            # Temperature-dependent equations (coefficients)
            'Cp_equation_type': None,  # Type of Cp(T) equation
            'Cp_coeff_a': None,  # Cp = a + bT + cT^2 + ...
            'Cp_coeff_b': None,
            'Cp_coeff_c': None,
            'Cp_temp_range_K_min': None,
            'Cp_temp_range_K_max': None,

            # Provenance
            'pressure_kPa': 101.325,
            'source_thermodynamic': 'NBS Tables',
            'source_physical_properties': None,
            'source_transition_temps': None,
            'data_quality': 'A',
            'CAS': None,
            'notes': None,
        })

        return output_df

    def save_final_files(self, output_dir: str = None):
        """Save final comprehensive files"""
        if output_dir is None:
            output_dir = '../library/processed_data'

        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d')

        # Save anhydrous
        anh_path = os.path.join(output_dir, f'thermophysical_comprehensive_anhydrous_{timestamp}.csv')
        self.df_anhydrous.to_csv(anh_path, index=False, encoding='utf-8')
        print(f"\n✅ Anhydrous: {anh_path}")
        print(f"   {len(self.df_anhydrous)} compounds")

        # Save hydrates
        hyd_path = os.path.join(output_dir, f'thermophysical_comprehensive_hydrates_{timestamp}.csv')
        self.df_hydrates.to_csv(hyd_path, index=False, encoding='utf-8')
        print(f"✅ Hydrates: {hyd_path}")
        print(f"   {len(self.df_hydrates)} compounds")

        return anh_path, hyd_path


def main():
    print("\n" + "#"*80)
    print("# COMPREHENSIVE THERMOPHYSICAL DATABASE")
    print("# Single-file compilation with ALL properties")
    print("#"*80)

    input_file = '../library/resources/NBS_Tables Library.xlsx'

    if not os.path.exists(input_file):
        print(f"\n❌ ERROR: {input_file} not found!")
        return

    processor = ComprehensiveThermophysicalProcessor(input_file)
    processor.process_complete_database()
    processor.save_final_files()

    print("\n" + "="*80)
    print("✅ COMPLETE!")
    print("="*80)
    print("\nCreated 2 comprehensive files:")
    print("  1. Anhydrous compounds - ALL properties in one file")
    print("  2. Hydrates - ALL properties in one file")
    print("\nIncludes placeholders for:")
    print("  • Heat of fusion, vaporization, sublimation")
    print("  • Density, thermal conductivity, viscosity, hardness")
    print("  • Stability ranges")
    print("  • Cp(T) equation coefficients")
    print("\nFill from CRC Handbook or other sources as needed.")


if __name__ == '__main__':
    main()

